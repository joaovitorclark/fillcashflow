
import pandas as pd
import yaml
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class FillcashFormatter:
    def __init__(self):
        self.base_path = Path(__file__).resolve().parent

    def run(self):
        print("▶️ Executando build e format...")
        self.future_card_bills = self.load_future_card_bills()
        self.build_cash_flow()
        self.format_sheet()

    def load_future_card_bills(self):
        path = self.base_path.parent / "statements" / "future_card_bills.xlsx"
        df = pd.read_excel(path)
        df["due_date"] = pd.to_datetime(df["due_date"]).dt.date
        df["last_digits"] = df["last_digits"].astype(str)  # ✅ correção aqui
        return df


    def build_cash_flow(self):
        statement_path = "outputs/current_account_statement.csv"
        config_path = "config.yml"
        output_path = "outputs/silver_statements.csv"

        today = datetime.today()
        start_date = datetime(today.year, 1, 1)
        end_date = datetime(today.year + 1, 12, 31)
        date_range = pd.date_range(start=start_date, end=end_date)

        df1 = pd.read_csv(statement_path, sep="|")
        df1["date"] = pd.to_datetime(df1["date"]).dt.date

        with open(config_path, "r") as f:
            config = yaml.safe_load(f)

        cards = config.get("cards", [])
        fixed_income = config.get("fixed_income", [])
        fixed_expenses = config.get("fixed_expenses", [])

        df = pd.DataFrame({"date": date_range})
        df["inflow"] = 0.0
        df["outflow"] = 0.0

        for card in cards:
            col = f"{card['bank'].capitalize()} - {card['name'].capitalize()} ({card['last_digits']})"
            df[col] = 0.0

        df["date"] = pd.to_datetime(df["date"]).dt.date

        # Aplicar faturas futuras de cartões diretamente nas colunas dos cartões
        for _, bill in self.future_card_bills.iterrows():
            col = f"{bill['bank'].capitalize()} - {bill['name'].capitalize()} ({bill['last_digits']})"
            if col in df.columns:
                df.loc[df['date'] == bill['due_date'], col] += bill['amount']

        extrato_por_dia = df1.groupby("date")[["inflow", "outflow"]].sum().reset_index()
        extrato_por_dia["date"] = pd.to_datetime(extrato_por_dia["date"]).dt.date

        dias_com_extrato = set(df1["date"].unique())
        last_real_date = df1["date"].max()

        for income in fixed_income:
            df.loc[
                (df["date"].apply(lambda d: d.day == income["day"] and d > last_real_date and d not in dias_com_extrato)),
                "inflow"
            ] += income["amount"]

        for expense in fixed_expenses:
            df.loc[
                (df["date"].apply(lambda d: d.day == expense["day"] and d > last_real_date and d not in dias_com_extrato)),
                "outflow"
            ] += expense["amount"]

        df_merge = df.merge(extrato_por_dia, on="date", how="left", suffixes=("", "_real"))
        df_merge["inflow"] = df_merge["inflow_real"].combine_first(df_merge["inflow"])
        df_merge["outflow"] = df_merge["outflow_real"].combine_first(df_merge["outflow"])
        df_merge = df_merge.drop(columns=["inflow_real", "outflow_real"])

        ordered_columns = ["date", "inflow", "outflow"] + [
            col for col in df_merge.columns if col not in ["date", "inflow", "outflow"]
        ]
        df_final = df_merge[ordered_columns]
        Path(output_path).parent.mkdir(exist_ok=True)
        df_final.to_csv(output_path, sep="|", index=False)
        print(f"✅ Silver statement saved to: {output_path}")

    def format_sheet(self):
        today_str = datetime.today().strftime("%d%m%y")
        output_file = f"outputs/format_sheet_{today_str}.xlsx"

        df, config = self.load_data("outputs/silver_statements.csv", "config.yml")
        Path("outputs").mkdir(parents=True, exist_ok=True)

        self.generate_cashflow_excel(df, config, output_file)
        print(f"✅ Cashflow file saved to: {output_file}")

    def load_data(self, csv_path: str, config_path: str):
        df = pd.read_csv(csv_path, sep="|")
        with open(config_path, "r") as f:
            config = yaml.safe_load(f)
        return df, config

    def generate_cashflow_excel(self, df, config, output_path):
        df["balance"] = 0
        wb = Workbook()
        ws = wb.active
        ws.title = "Cashflow"
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        header = [cell.value for cell in ws[1]]
        col_idx = {col: idx + 1 for idx, col in enumerate(header)}
        self.apply_card_styles(ws, col_idx, config["cards"])
        self.insert_balance_formulas(ws, col_idx, config["cards"])
        self.apply_conditional_formatting(ws, col_idx)
        self.center_all_cells(ws)
        wb.save(output_path)

    def apply_card_styles(self, ws, col_idx, cards):
        for card in cards:
            name = f"{card['bank'].capitalize()} - {card['name'].capitalize()} ({card['last_digits']})"
            colors = card["color"]
            if name in col_idx:
                col_letter = get_column_letter(col_idx[name])
                ws[f"{col_letter}1"].fill = PatternFill(start_color=colors[0][1:], end_color=colors[0][1:], fill_type="solid")
                for row in range(2, ws.max_row + 1):
                    fill_color = colors[1] if row % 2 == 0 else colors[2]
                    ws[f"{col_letter}{row}"].fill = PatternFill(start_color=fill_color[1:], end_color=fill_color[1:], fill_type="solid")

    def insert_balance_formulas(self, ws, col_idx, cards):
        balance_letter = get_column_letter(col_idx["balance"])
        inflow_letter = get_column_letter(col_idx["inflow"])
        outflow_letter = get_column_letter(col_idx["outflow"])
        date_letter = get_column_letter(col_idx["date"])

        for row in range(2, ws.max_row + 1):
            current_date = ws[f"{date_letter}{row}"].value
            if isinstance(current_date, str):
                current_date = datetime.strptime(current_date, "%Y-%m-%d").date()
            future_charges = self.future_card_bills[self.future_card_bills['due_date'] == current_date]

            deductions = []
            for _, bill in future_charges.iterrows():
                for col in col_idx:
                    if bill['bank'].capitalize() in col and bill['last_digits'] in col:
                        card_letter = get_column_letter(col_idx[col])
                        deductions.append(f"{card_letter}{row}")

            inflow_cell = f"{inflow_letter}{row}"
            outflow_cell = f"{outflow_letter}{row}"
            prev_balance = f"{balance_letter}{row - 1}" if row > 2 else "0"
            deduction_expr = "-(" + "+".join(deductions) + ")" if deductions else ""
            formula = f"={prev_balance}+{inflow_cell}-{outflow_cell}{deduction_expr}"

            cell = ws[f"{balance_letter}{row}"]
            cell.value = formula
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def apply_conditional_formatting(self, ws, col_idx):
        balance_letter = get_column_letter(col_idx["balance"])
        balance_range = f"{balance_letter}2:{balance_letter}{ws.max_row}"
        rule = ColorScaleRule(
            start_type='num', start_value=-5000, start_color='FF0000',
            mid_type='num', mid_value=0, mid_color='FFFF00',
            end_type='num', end_value=20000, end_color='00FF00'
        )
        ws.conditional_formatting.add(balance_range, rule)
        ws[f"{balance_letter}1"].font = Font(bold=True)

    def center_all_cells(self, ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
